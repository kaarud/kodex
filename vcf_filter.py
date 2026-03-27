"""vcf_filter.py — Filter NiourK XLSX for TSC1/TSC2 pathogenic candidates."""

import argparse
import hashlib
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import yaml
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Load & normalise
# ---------------------------------------------------------------------------

def load_variants(xlsx_path: Path, genes: list[str] | None = None) -> pd.DataFrame:
    """Load the 'Variants' sheet and strip NiourK sample-prefix from columns.

    NiourK format: per-sample columns are named '<sample_id>\\n<column_name>'.
    We detect the sample prefix and strip it so downstream code uses clean names.
    """
    df = pd.read_excel(xlsx_path, sheet_name="Variants", engine="openpyxl")

    # Detect sample prefix: look for columns containing '\n'
    renamed = {}
    for col in df.columns:
        if "\n" in str(col):
            # Take everything after the first newline
            clean = str(col).split("\n", 1)[1]
            renamed[col] = clean
    df = df.rename(columns=renamed)

    # Filter to target genes
    if genes:
        df = df[df["SYMBOL"].isin(genes)].reset_index(drop=True)

    # Deduplicate by REF_TRANS: for each position, keep only the canonical
    # transcript (REF_TRANS == "VRAI") when one exists; keep all rows otherwise.
    pos_key = ["CHR", "POS", "REF", "ALT"]
    if "REF_TRANS" in df.columns and all(c in df.columns for c in pos_key):
        is_ref = df["REF_TRANS"].astype(str).str.upper().isin(["VRAI", "TRUE", "1"])
        has_ref = df.groupby(pos_key)["REF_TRANS"].transform(
            lambda s: s.astype(str).str.upper().isin(["VRAI", "TRUE", "1"]).any()
        ).astype(bool)
        df = df[~has_ref | is_ref].reset_index(drop=True)

    return df


def load_config(config_path: Path) -> dict:
    """Load YAML config file."""
    with open(config_path) as f:
        return yaml.safe_load(f)


# ---------------------------------------------------------------------------
# 2. Exclusion filters
# ---------------------------------------------------------------------------

def _safe_gt(series: pd.Series, threshold, fill=0.0) -> pd.Series:
    """Compare series > threshold, treating NaN as `fill`."""
    return series.fillna(fill) > threshold


def apply_exclusions(df: pd.DataFrame, cfg: dict) -> tuple[pd.DataFrame, dict]:
    """Apply sequential exclusion filters. Returns (filtered_df, report_dict).

    report_dict maps filter name → count of rows removed at that step.
    Filters are applied in spec order; counts are cumulative (each step
    operates on the survivors of the previous step).
    """
    exc = cfg["exclusions"]
    report = {}

    # 1. gnomAD > max_af (NaN = novel → keep)
    before = len(df)
    mask = _safe_gt(df["MAX_AF_GNOMADEX2.1"], exc["gnomad_max_af"], fill=0.0)
    df = df[~mask].reset_index(drop=True)
    report["gnomad"] = before - len(df)

    # 2. RNCNT_TOTAL > max
    before = len(df)
    mask = _safe_gt(df["RNCNT_TOTAL"], exc["rncnt_max_total"])
    df = df[~mask].reset_index(drop=True)
    report["rncnt_total"] = before - len(df)

    # 3. PJCNT_TOTAL > max
    before = len(df)
    mask = _safe_gt(df["PJCNT_TOTAL"], exc["pjcnt_max_total"])
    df = df[~mask].reset_index(drop=True)
    report["pjcnt_total"] = before - len(df)

    # 4. RNCNT_HOM > 0 (NaN → 0 → keep)
    if exc.get("exclude_hom", True):
        before = len(df)
        mask = _safe_gt(df["RNCNT_HOM"], 0)
        df = df[~mask].reset_index(drop=True)
        report["hom"] = before - len(df)

    # 5. ClinVar Benign / Likely_benign
    if exc.get("exclude_clinvar_benign", True):
        before = len(df)
        clinvar = df["CLINVAR"].fillna("").astype(str)
        mask = clinvar.str.contains("Benign|Likely_benign", case=False, regex=True)
        df = df[~mask].reset_index(drop=True)
        report["clinvar_benign"] = before - len(df)

    # 6. Artifact: recurrent mosaic hotspot
    art_mos = exc.get("mosaic_artifact", {})
    if art_mos.get("enabled", False):
        before = len(df)
        mask = (
            (_safe_gt(df["RNCNT_LOW"], art_mos["rncnt_low_max"]))
            & (df["AF_PCT"].fillna(0) < art_mos["af_max_pct"])
        )
        df = df[~mask].reset_index(drop=True)
        report["mosaic_artifact"] = before - len(df)

    return df, report


# ---------------------------------------------------------------------------
# 3. Tier classification
# ---------------------------------------------------------------------------

def classify_sheets(df: pd.DataFrame, cfg: dict) -> dict[str, pd.DataFrame]:
    """Classify surviving variants into named sheets (fiche d'analyse STB).

    Each sheet is independent — a variant can appear in multiple sheets.
    Returns dict of sheet_name → DataFrame.
    """
    scfg = cfg["sheets"]
    sheets = {}

    clinvar = df["CLINVAR"].fillna("").astype(str)

    # 1. HIGH_IMPACT
    if scfg.get("high_impact", {}).get("enabled", False):
        sheets["HIGH_IMPACT"] = df[df["IMPACT"] == "HIGH"].copy()

    # 2. CLINVAR_P_LP (exclude Conflicting entries)
    if scfg.get("clinvar_p_lp", {}).get("enabled", False):
        mask = (
            clinvar.str.contains("Pathogenic|Likely_pathogenic", case=False, regex=True)
            & ~clinvar.str.contains("Conflicting", case=False)
        )
        sheets["CLINVAR_P_LP"] = df[mask].copy()

    # 3. CLINVAR_CONFLICTING
    if scfg.get("clinvar_conflicting", {}).get("enabled", False):
        mask = clinvar.str.contains("Conflicting", case=False)
        sheets["CLINVAR_CONFLICTING"] = df[mask].copy()

    # 4. CLINVAR_VUS
    if scfg.get("clinvar_vus", {}).get("enabled", False):
        mask = clinvar.str.contains("Uncertain_significance", case=False)
        sheets["CLINVAR_VUS"] = df[mask].copy()

    # 5. CLINVAR_NOT_PROVIDED
    if scfg.get("clinvar_not_provided", {}).get("enabled", False):
        mask = clinvar.str.contains("not_provided", case=False)
        sheets["CLINVAR_NOT_PROVIDED"] = df[mask].copy()

    # 6. FQ_MODERATE (gnomAD < 0.1% + MODERATE)
    if scfg.get("fq_moderate", {}).get("enabled", False):
        fq_cfg = scfg["fq_moderate"]
        gnomad_ok = df["MAX_AF_GNOMADEX2.1"].fillna(0) < fq_cfg["gnomad_max_af"]
        mask = gnomad_ok & (df["IMPACT"] == "MODERATE")
        sheets["FQ_MODERATE"] = df[mask].copy()

    # 7. FQ_LOW (gnomAD < 0.1% + LOW)
    if scfg.get("fq_low", {}).get("enabled", False):
        fq_cfg = scfg["fq_low"]
        gnomad_ok = df["MAX_AF_GNOMADEX2.1"].fillna(0) < fq_cfg["gnomad_max_af"]
        mask = gnomad_ok & (df["IMPACT"] == "LOW")
        sheets["FQ_LOW"] = df[mask].copy()

    # 8. FQ_MODIFIER (gnomAD < 0.1% + MODIFIER + spliceAI HIGH OR SPiP)
    if scfg.get("fq_modifier", {}).get("enabled", False):
        fq_cfg = scfg["fq_modifier"]
        gnomad_ok = df["MAX_AF_GNOMADEX2.1"].fillna(0) < fq_cfg["gnomad_max_af"]
        spliceai = df["SpliceAI_num"].fillna(0)
        spip_int = df["SPiP_Interpretation"].fillna("").astype(str)
        splice_ok = (spliceai >= fq_cfg["spliceai_threshold"]) | (spip_int == "Altered")
        mask = gnomad_ok & (df["IMPACT"] == "MODIFIER") & splice_ok
        sheets["FQ_MODIFIER"] = df[mask].copy()

    # 9. 4_CALLERS
    if scfg.get("four_callers", {}).get("enabled", False):
        fc_cfg = scfg["four_callers"]
        mask = df["CALLNB"].fillna(0) >= fc_cfg["min_callers"]
        sheets["4_CALLERS"] = df[mask].copy()

    # 10. LOVD
    if scfg.get("lovd", {}).get("enabled", False):
        lovd = df["LOVD"].fillna("").astype(str)
        mask = (lovd != "") & (lovd != ".")
        sheets["LOVD"] = df[mask].copy()

    # 11. MOSAIC (5% < VAF ≤ 35%, RNCNT_TOTAL ≤ 2)
    if scfg.get("mosaic", {}).get("enabled", False):
        mo_cfg = scfg["mosaic"]
        af = df["AF_PCT"].fillna(0)
        rncnt = df["RNCNT_TOTAL"].fillna(0)
        mask = (
            (af > mo_cfg["af_min_pct"])
            & (af <= mo_cfg["af_max_pct"])
            & (rncnt <= mo_cfg["rncnt_max_total"])
        )
        sheets["MOSAIC"] = df[mask].copy()

    # 12. DEEP_MOSAIC (0.5 % < VAF ≤ 5 %, RNCNT_TOTAL ≤ 2, DP ≥ min_dp)
    # Captures very deep somatic mosaics missed by the 5 % MOSAIC floor.
    if scfg.get("deep_mosaic", {}).get("enabled", False):
        dm_cfg = scfg["deep_mosaic"]
        af = df["AF_PCT"].fillna(0)
        rncnt = df["RNCNT_TOTAL"].fillna(0)
        dp = df["DP"].fillna(0)
        # SB est un ratio 0–1 : 0 = tous reads sur brin–, 1 = tous reads sur brin+
        # Les deux extrêmes indiquent un biais total de brin (artefact probable)
        if "SB" in df.columns and dm_cfg.get("sb_exclude_extremes", True):
            sb = pd.to_numeric(df["SB"], errors="coerce")
            sb_ok = sb.isna() | ((sb > 0) & (sb < 1))
        else:
            sb_ok = pd.Series(True, index=df.index)
        mask = (
            (af > dm_cfg["af_min_pct"])
            & (af <= dm_cfg["af_max_pct"])
            & (rncnt <= dm_cfg["rncnt_max_total"])
            & (dp >= dm_cfg["min_dp"])
            & sb_ok
        )
        sheets["DEEP_MOSAIC"] = df[mask].copy()

    return sheets


# ---------------------------------------------------------------------------
# 4. Pre-report
# ---------------------------------------------------------------------------

def build_pre_report(
    total_rows: int,
    unique_variants: int,
    exclusion_report: dict,
    sheets: dict[str, pd.DataFrame],
    config_path: str,
    input_name: str,
) -> str:
    """Build the pre-report string for terminal output."""
    lines = []
    today = date.today().isoformat()
    lines.append(f"Pre-report {input_name} — {today}")
    lines.append(f"config : {config_path}")
    lines.append("")
    lines.append(f"Variants TSC1/TSC2 bruts (lignes)         : {total_rows}")
    lines.append(f"  dont variants uniques (CHR+POS+REF+ALT) : {unique_variants}")
    lines.append("")
    lines.append("--- Exclusions communes (séquentielles, sur lignes) ---")

    labels = {
        "gnomad": "gnomAD > 0.5 %",
        "rncnt_total": "RNCNT_TOTAL > seuil",
        "pjcnt_total": "PJCNT_TOTAL > seuil",
        "hom": "RNCNT_HOM > 0",
        "clinvar_benign": "ClinVar Benign",
        "mosaic_artifact": "Artefact mosaïque récurrente",
    }

    remaining = total_rows
    for key, label in labels.items():
        count = exclusion_report.get(key, 0)
        remaining -= count
        lines.append(f"  {label:<36s}: -{count:>4d}  → {remaining:>4d} restants")

    lines.append("")
    lines.append(f"--- Onglets fiche STB (sur {remaining} lignes restantes, indépendants) ---")
    for name, sheet_df in sheets.items():
        n_rows = len(sheet_df)
        n_unique = len(sheet_df.drop_duplicates(subset=["CHR", "POS", "REF", "ALT"]))
        lines.append(f"  {name:<24s}: {n_unique:>3d} variants ({n_rows} lignes)")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 5. Excel output
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")   # bleu NiourK
_HEADER_FONT = Font(bold=True, size=10)
_DATA_FONT   = Font(size=10)
_NO_WRAP     = Alignment(wrap_text=False, vertical="top")
_HEADER_ALIGN = Alignment(wrap_text=False, vertical="center", horizontal="center")
_ROW_HEIGHT  = 15   # pt — hauteur fixe données
_BLUE_FONT   = Font(color="0000FF", size=10)  # doublons HGVSc
_BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")  # downstream → gène voisin

# Gènes voisins de TSC1/TSC2 — coordonnées GRCh38 (région étendue)
# TSC2 finit ~2 089 490 → PKD1 commence ~2 138 711.
# On couvre l'intergénique TSC2-PKD1 + PKD1 entier.
_NEIGHBOR_GENES = {
    "PKD1": {"chr": "chr16", "start": 2089491, "end": 2185899},
    "RHEB": {"chr": "chr7",  "start": 151497637, "end": 151553587},
}

# Conditional colour palette
_C_GREEN  = PatternFill("solid", fgColor="C6EFCE")  # bon / favorable
_C_LGREEN = PatternFill("solid", fgColor="ABEBC6")  # 3 callers / rare
_C_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # attention / modéré
_C_ORANGE = PatternFill("solid", fgColor="FAD7A0")  # alerte
_C_RED    = PatternFill("solid", fgColor="FFC7CE")  # mauvais / pathogène fort


def _num(v):
    """Safe float conversion; returns None on failure."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_fill(col: str, value) -> "PatternFill | None":
    """Return a conditional fill for (column, value), or None if no rule applies."""
    v = _num(value)
    s = str(value).strip() if value not in (None, float("nan")) else ""

    if col == "SpliceAI_num":
        if v is None: return None
        if v >= 0.8:  return _C_RED
        if v >= 0.5:  return _C_YELLOW

    elif col == "CADD_PHRED":
        if v is None: return None
        if v >= 30:   return _C_RED
        if v >= 20:   return _C_YELLOW

    elif col == "REVEL":
        if v is None: return None
        if v >= 0.75: return _C_RED
        if v >= 0.5:  return _C_YELLOW
        if v < 0.5:   return _C_GREEN

    elif col == "AF_PCT":
        if v is None: return None
        if 40 <= v <= 60:  return _C_GREEN    # hétérozygote germinal
        if 5  <  v <  40:  return _C_YELLOW   # zone mosaïque
        if 0.5 < v <= 5:   return _C_ORANGE   # mosaïque profond

    elif col == "SB":
        if v is None: return None
        if 0.4 <= v <= 0.6:              return _C_GREEN
        if 0.3 <= v < 0.4 or 0.6 < v <= 0.7: return _C_YELLOW
        if v < 0.3 or v > 0.7:          return _C_RED

    elif col == "CALLNB":
        if v is None: return None
        if v >= 4:    return _C_GREEN
        if v == 3:    return _C_LGREEN
        if v == 2:    return _C_YELLOW
        if v == 1:    return _C_ORANGE
        if v == 0:    return _C_RED

    elif col == "MAX_AF_GNOMADEX2.1":
        if v is None or v == 0: return _C_GREEN   # absent gnomAD → novel
        if v > 0.001:           return _C_YELLOW  # > 0.1 % → commun

    elif col == "RNCNT_TOTAL":
        if v is None: return None
        if v == 0:            return _C_GREEN
        if 1 <= v <= 2:       return _C_LGREEN
        if 3 <= v <= 4:       return _C_YELLOW

    elif col == "PJCNT_TOTAL":
        if v is None: return None
        if v == 0:            return _C_GREEN
        if 1 <= v <= 5:       return _C_LGREEN
        if 6 <= v <= 10:      return _C_YELLOW

    elif col == "SPiP_Interpretation":
        if s == "Altered":    return _C_RED
        if s == "Conserved":  return _C_GREEN

    elif col == "IMPACT":
        if s == "HIGH":       return _C_RED
        if s == "MODERATE":   return _C_YELLOW
        if s == "LOW":        return _C_LGREEN

    elif col == "USER_CLASS":
        if s in ("5",):       return _C_RED
        if s in ("4",):       return _C_ORANGE
        if s in ("3",):       return _C_YELLOW
        if s in ("2", "1"):   return _C_GREEN

    return None


# ---------------------------------------------------------------------------
# 5b. Variant quality scoring — couleur RGB 3 canaux sur cellule Import DefGen
#
#   R = 1 − qualité_technique  (SB, CALLNB, DP/AF)
#   G = signal_biologique      (ClinVar, IMPACT, CADD, REVEL, SpliceAI)
#   B = rareté                 (RNCNT, PJCNT, gnomAD, LOVD)
#
#   Cyan  #00FFFF : idéal       | Vert  #00FF00 : bio forte, non rare
#   Jaune #FFFF00 : faux-positif dangereux (bio + artefact)
#   Rouge #FF0000 : artefact pur | Noir #000000 : sans intérêt
#
#   Règle spéciale : absent gnomAD + PJCNT ≥ seuil → contribution gnomAD = 0
#   (artefact panel probable, pas un variant rare)
# ---------------------------------------------------------------------------

# Default scoring config — overridden by config.yaml "scoring" key
_DEFAULT_SCORING = {
    "enabled": True,
    "verdicts": {
        "high": {"threshold": 0.7, "label": "Variant solide, à valider"},
        "mixed": {"threshold": 0.5, "label": "À évaluer — signaux mixtes"},
        "low": {"threshold": 0.3, "label": "Probablement artefact ou inintéressant"},
        "artifact": {"threshold": 0.0, "label": "Artefact probable"},
    },
    "technical": {
        "sb": {
            "weight": 2.0,
            "good_min": 0.3,
            "good_max": 0.7,
            "warn_min": 0.2,
            "warn_max": 0.8,
        },
        "callnb": {
            "weight": 2.0,
            "excellent": 4,
            "good": 3,
            "moderate": 2,
        },
        "dp_af": {
            "weight": 0.5,
            "min_alt_reads": 5,
        },
    },
    "biological": {
        "clinvar": {
            "weight": 2.5,
            "pathogenic": 1.0,
            "conflicting": 0.4,
            "vus": 0.6,
            "benign": 0.0,
            "absent": 0.4,
            "other": 0.2,
        },
        "impact": {
            "weight": 2.0,
            "HIGH": 1.0,
            "MODERATE": 0.75,
            "LOW": 0.25,
            "MODIFIER": 0.0,
        },
        "cadd": {
            "weight": 1.0,
            "good": 30,
            "moderate": 20,
        },
        "revel": {
            "weight": 1.0,
            "good": 0.75,
            "moderate": 0.5,
        },
        "spliceai": {
            "weight": 1.0,
            "good": 0.8,
            "moderate": 0.5,
        },
    },
    "rarity": {
        "rncnt": {
            "weight": 1.5,
            "good_max": 1,          # 0 ou 1 = patient unique dans la base
            "moderate_max": 3,
            "warn_max": 5,
        },
        "pjcnt": {
            "weight": 1.0,
            "good_max": 0,
            "moderate_max": 5,
        },
        "gnomad": {
            "weight": 1.5,
            "ultra_rare": 0.0001,
            "rare": 0.001,
        },
        "lovd": {
            "weight": 0.5,
        },
    },
}


def _deep_merge(base: dict, override: dict) -> dict:
    """Recursively merge override into base, returning a new dict."""
    result = base.copy()
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = _deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def _scores_to_rgb(bio: float, tech_quality: float, rarity: float) -> str:
    """Map three 0–1 axis scores to an RGB hex color.

    R = artifact risk  (1 − tech_quality): high → red
    G = biological significance           : high → green
    B = variant rarity                    : high → blue

    Key color signatures:
      Cyan   (G+B, R=0) : ideal candidate — biology + rare + clean sequencing
      Yellow (R+G, B=0) : dangerous false positive — biology + artifact
      Red    (R only)   : artifact, no biological signal
      Black  (all low)  : uninteresting but technically clean
    """
    r = int((1.0 - max(0.0, min(1.0, tech_quality))) * 255)
    g = int(max(0.0, min(1.0, bio)) * 255)
    b = int(max(0.0, min(1.0, rarity)) * 255)
    return f"{r:02X}{g:02X}{b:02X}"


def _score_variant(row: dict, scoring: dict | None = None) -> tuple[float, float, float, float, str]:
    """Compute per-axis quality scores (0–1) and a detailed comment.

    scoring: merged config dict (defaults + user overrides from config.yaml).
    Returns (bio_norm, tech_norm, rarity_norm, combined_norm, comment_text).

    Axes:
      bio_norm      → biological significance (ClinVar, IMPACT, CADD, REVEL, SpliceAI)
      tech_norm     → technical quality (SB, CALLNB, DP/AF) — high = clean sequencing
      rarity_norm   → variant rarity (gnomAD, RNCNT, PJCNT, LOVD)
      combined_norm → weighted average across all axes (used for verdict text)
    """
    cfg = _deep_merge(_DEFAULT_SCORING, scoring or {})
    tech_cfg = cfg["technical"]
    bio_cfg = cfg["biological"]
    rarity_cfg = cfg["rarity"]

    tech_pts = 0.0
    tech_max = 0.0
    bio_pts = 0.0
    bio_max = 0.0
    rarity_pts = 0.0
    rarity_max = 0.0

    tech_lines: list[str] = []
    bio_lines: list[str] = []
    rarity_lines: list[str] = []

    # --- Technical axis (SB, CALLNB, DP/AF) ---

    # SB (strand bias)
    sb_cfg = tech_cfg["sb"]
    sb = _num(row.get("SB"))
    if sb is not None:
        w = sb_cfg["weight"]
        tech_max += w
        if sb_cfg["good_min"] <= sb <= sb_cfg["good_max"]:
            tech_pts += w
            tech_lines.append(f"  ✓ SB={sb:.2f} (équilibré)")
        elif sb_cfg["warn_min"] <= sb <= sb_cfg["warn_max"]:
            tech_pts += w * 0.5
            tech_lines.append(f"  ⚠ SB={sb:.2f} (léger biais de brin)")
        else:
            tech_lines.append(f"  ✗ SB={sb:.2f} (biais de brin fort)")

    # CALLNB
    cn_cfg = tech_cfg["callnb"]
    callnb = _num(row.get("CALLNB"))
    if callnb is not None:
        w = cn_cfg["weight"]
        tech_max += w
        if callnb >= cn_cfg["excellent"]:
            tech_pts += w
            tech_lines.append(f"  ✓ CALLNB={int(callnb)} ({int(callnb)} callers concordants)")
        elif callnb >= cn_cfg["good"]:
            tech_pts += w * 0.75
            tech_lines.append(f"  ✓ CALLNB={int(callnb)} ({int(callnb)} callers concordants)")
        elif callnb >= cn_cfg["moderate"]:
            tech_pts += w * 0.5
            tech_lines.append(f"  ⚠ CALLNB={int(callnb)} ({int(callnb)} callers seulement)")
        else:
            tech_lines.append(f"  ✗ CALLNB={int(callnb)} (caller unique)")

    # DP cohérence avec AF
    da_cfg = tech_cfg["dp_af"]
    dp = _num(row.get("DP"))
    af = _num(row.get("AF_PCT"))
    if dp is not None and af is not None:
        w = da_cfg["weight"]
        tech_max += w
        expected_alt_reads = dp * af / 100
        if expected_alt_reads >= da_cfg["min_alt_reads"]:
            tech_pts += w
            tech_lines.append(f"  ✓ DP={int(dp)}, AF={af:.1f}% (~{int(expected_alt_reads)} reads alt)")
        else:
            tech_lines.append(f"  ⚠ DP={int(dp)}, AF={af:.1f}% (~{int(expected_alt_reads)} reads alt — faible)")

    # --- Biological axis (ClinVar, IMPACT, CADD, REVEL, SpliceAI) ---

    # ClinVar
    cv_cfg = bio_cfg["clinvar"]
    clinvar = str(row.get("CLINVAR") or "").strip()
    w = cv_cfg["weight"]
    bio_max += w
    if clinvar and clinvar not in ("nan", "None"):
        cl = clinvar.lower()
        if "pathogenic" in cl and "conflicting" not in cl and "benign" not in cl:
            bio_pts += w * cv_cfg["pathogenic"]
            bio_lines.append(f"  ✓ ClinVar: {clinvar}")
        elif "conflicting" in cl:
            bio_pts += w * cv_cfg["conflicting"]
            bio_lines.append(f"  ⚠ ClinVar: {clinvar}")
        elif "uncertain" in cl:
            bio_pts += w * cv_cfg["vus"]
            bio_lines.append(f"  ~ ClinVar: VUS")
        elif "benign" in cl:
            bio_pts += w * cv_cfg["benign"]
            bio_lines.append(f"  ✗ ClinVar: {clinvar}")
        else:
            bio_pts += w * cv_cfg["other"]
            bio_lines.append(f"  ~ ClinVar: {clinvar}")
    else:
        bio_pts += w * cv_cfg["absent"]
        bio_lines.append("  ~ ClinVar: absent")

    # IMPACT
    im_cfg = bio_cfg["impact"]
    impact = str(row.get("IMPACT") or "").strip()
    w = im_cfg["weight"]
    bio_max += w
    ratio = im_cfg.get(impact, im_cfg.get("MODIFIER", 0.0))
    bio_pts += w * ratio
    consequence = row.get("Consequence", "")
    if ratio >= 0.75:
        bio_lines.append(f"  ✓ IMPACT: {impact} ({consequence})")
    elif ratio >= 0.5:
        bio_lines.append(f"  ~ IMPACT: {impact} ({consequence})")
    elif ratio > 0:
        bio_lines.append(f"  ⚠ IMPACT: {impact} ({consequence})")
    else:
        bio_lines.append(f"  ✗ IMPACT: {impact or 'MODIFIER'}")

    # CADD
    ca_cfg = bio_cfg["cadd"]
    cadd = _num(row.get("CADD_PHRED"))
    if cadd is not None:
        w = ca_cfg["weight"]
        bio_max += w
        if cadd >= ca_cfg["good"]:
            bio_pts += w
            bio_lines.append(f"  ✓ CADD={cadd:.1f}")
        elif cadd >= ca_cfg["moderate"]:
            bio_pts += w * 0.5
            bio_lines.append(f"  ~ CADD={cadd:.1f}")
        else:
            bio_lines.append(f"  ✗ CADD={cadd:.1f}")

    # REVEL
    re_cfg = bio_cfg["revel"]
    revel = _num(row.get("REVEL"))
    if revel is not None:
        w = re_cfg["weight"]
        bio_max += w
        if revel >= re_cfg["good"]:
            bio_pts += w
            bio_lines.append(f"  ✓ REVEL={revel:.2f}")
        elif revel >= re_cfg["moderate"]:
            bio_pts += w * 0.5
            bio_lines.append(f"  ~ REVEL={revel:.2f}")
        else:
            bio_lines.append(f"  ✗ REVEL={revel:.2f}")

    # SpliceAI
    sp_cfg = bio_cfg["spliceai"]
    spliceai = _num(row.get("SpliceAI_num"))
    if spliceai is not None:
        w = sp_cfg["weight"]
        bio_max += w
        if spliceai >= sp_cfg["good"]:
            bio_pts += w
            bio_lines.append(f"  ✓ SpliceAI={spliceai:.2f}")
        elif spliceai >= sp_cfg["moderate"]:
            bio_pts += w * 0.5
            bio_lines.append(f"  ~ SpliceAI={spliceai:.2f}")
        else:
            bio_lines.append(f"  ✗ SpliceAI={spliceai:.2f}")

    # --- Rarity axis (RNCNT, PJCNT, gnomAD, LOVD) ---

    # RNCNT_TOTAL
    rn_cfg = rarity_cfg["rncnt"]
    rncnt = _num(row.get("RNCNT_TOTAL"))
    if rncnt is not None:
        w = rn_cfg["weight"]
        rarity_max += w
        if rncnt <= rn_cfg["good_max"]:
            rarity_pts += w
            rarity_lines.append(f"  ✓ RNCNT={int(rncnt)} (patient unique dans la base)")
        elif rncnt <= rn_cfg["moderate_max"]:
            rarity_pts += w * 0.67
            rarity_lines.append(f"  ~ RNCNT={int(rncnt)} (peu fréquent)")
        elif rncnt <= rn_cfg["warn_max"]:
            rarity_pts += w * 0.33
            rarity_lines.append(f"  ⚠ RNCNT={int(rncnt)} (récurrent dans le run)")
        else:
            rarity_lines.append(f"  ✗ RNCNT={int(rncnt)} (très récurrent)")

    # PJCNT_TOTAL
    pj_cfg = rarity_cfg["pjcnt"]
    pjcnt = _num(row.get("PJCNT_TOTAL"))
    if pjcnt is not None:
        w = pj_cfg["weight"]
        rarity_max += w
        if pjcnt <= pj_cfg["good_max"]:
            rarity_pts += w
            rarity_lines.append(f"  ✓ PJCNT={int(pjcnt)} (jamais vu inter-projets)")
        elif pjcnt <= pj_cfg["moderate_max"]:
            rarity_pts += w * 0.5
            rarity_lines.append(f"  ~ PJCNT={int(pjcnt)}")
        else:
            rarity_lines.append(f"  ⚠ PJCNT={int(pjcnt)} (fréquent inter-projets)")

    # gnomAD — avec détection du paradoxe "absent gnomAD + récurrent dans nos runs"
    # Si absent gnomAD mais PJCNT élevé → probable artefact panel (non vu en population car technique,
    # pas car rare). Dans ce cas l'absence gnomAD ne vaut rien comme argument de rareté.
    gn_cfg = rarity_cfg["gnomad"]
    gnomad = _num(row.get("MAX_AF_GNOMADEX2.1"))
    w = gn_cfg["weight"]
    rarity_max += w
    _pjcnt_artifact_threshold = rarity_cfg.get("gnomad_pjcnt_benign_threshold", 10)
    _gnomad_absent = gnomad is None or gnomad == 0
    _pjcnt_high = pjcnt is not None and pjcnt >= _pjcnt_artifact_threshold
    if _gnomad_absent and _pjcnt_high:
        # Score nul : l'absence gnomAD n'est pas un argument de rareté ici
        rarity_lines.append(
            f"  ✗ Absent gnomAD mais PJCNT={int(pjcnt)} — artefact panel probable"
            f" (non représenté en population car technique, pas car rare)"
        )
    elif _gnomad_absent:
        rarity_pts += w
        rarity_lines.append("  ✓ Absent gnomAD (novel)")
    elif gnomad < gn_cfg["ultra_rare"]:
        rarity_pts += w * 0.67
        rarity_lines.append(f"  ✓ gnomAD={gnomad:.5f} (ultra-rare)")
    elif gnomad < gn_cfg["rare"]:
        rarity_pts += w * 0.33
        rarity_lines.append(f"  ~ gnomAD={gnomad:.4f} (rare)")
    else:
        rarity_lines.append(f"  ⚠ gnomAD={gnomad:.4f} (fréquent)")

    # LOVD
    lv_cfg = rarity_cfg["lovd"]
    lovd = str(row.get("LOVD") or "").strip()
    if lovd and lovd not in ("", ".", "nan", "None"):
        w = lv_cfg["weight"]
        rarity_max += w
        rarity_pts += w
        rarity_lines.append("  ✓ Présent dans LOVD")

    # --- Sub-score display counts ---
    tech_good = sum(1 for l in tech_lines if "✓" in l)
    tech_total = len(tech_lines)
    bio_good = sum(1 for l in bio_lines if "✓" in l)
    bio_total = len(bio_lines)
    rarity_good = sum(1 for l in rarity_lines if "✓" in l)
    rarity_total = len(rarity_lines)

    # --- Normalized axis scores ---
    bio_norm = bio_pts / bio_max if bio_max > 0 else 0.5
    tech_norm = tech_pts / tech_max if tech_max > 0 else 0.5
    rarity_norm = rarity_pts / rarity_max if rarity_max > 0 else 0.5
    all_pts = bio_pts + tech_pts + rarity_pts
    all_max = bio_max + tech_max + rarity_max
    combined_norm = all_pts / all_max if all_max > 0 else 0.5

    # --- Build quick overview ---
    pros = []
    cons = []

    if sb is not None:
        if sb_cfg["good_min"] <= sb <= sb_cfg["good_max"]:
            pros.append("SB équilibré")
        elif sb < sb_cfg["warn_min"] or sb > sb_cfg["warn_max"]:
            cons.append(f"SB={sb:.2f} biais fort")
        else:
            cons.append(f"SB={sb:.2f} léger biais")

    if callnb is not None:
        if callnb >= cn_cfg["excellent"]:
            pros.append(f"{int(callnb)} callers")
        elif callnb >= cn_cfg["good"]:
            pros.append(f"{int(callnb)} callers")
        elif callnb <= 1:
            cons.append("1 seul caller")
        else:
            cons.append(f"{int(callnb)} callers seulement")

    if rncnt is not None:
        if rncnt <= rn_cfg["good_max"]:
            pros.append("unique run")
        elif rncnt > rn_cfg["warn_max"]:
            cons.append(f"RNCNT={int(rncnt)} récurrent")

    if pjcnt is not None and pjcnt > pj_cfg["moderate_max"]:
        cons.append(f"PJCNT={int(pjcnt)} fréquent inter-projets")

    if dp is not None and af is not None:
        alt_reads = dp * af / 100
        if alt_reads < da_cfg["min_alt_reads"]:
            cons.append(f"~{int(alt_reads)} reads alt (faible)")

    # Biological highlights
    if clinvar and clinvar not in ("nan", "None"):
        cl = clinvar.lower()
        if "pathogenic" in cl and "conflicting" not in cl and "benign" not in cl:
            pros.append("ClinVar P/LP")
        elif "conflicting" in cl:
            cons.append("ClinVar conflicting")
        elif "benign" in cl:
            cons.append("ClinVar Benign")

    if impact == "HIGH":
        pros.append(f"HIGH impact ({consequence})")
    elif impact == "MODIFIER":
        cons.append("MODIFIER")

    if cadd is not None:
        if cadd >= ca_cfg["good"]:
            pros.append(f"CADD={cadd:.0f}")
        elif cadd < ca_cfg["moderate"]:
            cons.append(f"CADD={cadd:.0f} faible")

    if revel is not None:
        if revel >= re_cfg["good"]:
            pros.append(f"REVEL={revel:.2f}")
        elif revel < re_cfg["moderate"]:
            cons.append(f"REVEL={revel:.2f} faible")

    if spliceai is not None and spliceai >= sp_cfg["good"]:
        pros.append(f"SpliceAI={spliceai:.2f}")

    if _gnomad_absent and _pjcnt_high:
        cons.append(f"absent gnomAD mais PJCNT={int(pjcnt)} → artefact panel")
    elif _gnomad_absent:
        pros.append("absent gnomAD")
    elif gnomad is not None and gnomad >= gn_cfg["rare"]:
        cons.append(f"gnomAD={gnomad:.4f}")

    if lovd and lovd not in ("", ".", "nan", "None"):
        pros.append("LOVD")

    # --- Build verdict ---
    verdicts = cfg["verdicts"]
    if combined_norm >= verdicts["high"]["threshold"]:
        verdict = verdicts["high"]["label"]
    elif combined_norm >= verdicts["mixed"]["threshold"]:
        verdict = verdicts["mixed"]["label"]
    elif combined_norm >= verdicts["low"]["threshold"]:
        verdict = verdicts["low"]["label"]
    else:
        verdict = verdicts["artifact"]["label"]

    # --- Assemble comment ---
    lines = []

    lines.append(f"{'=' * 35}")
    lines.append(f"  SCORE : {combined_norm:.0%} — {verdict}")
    lines.append(f"  Tech {tech_good}/{tech_total}  |  Bio {bio_good}/{bio_total}  |  Rareté {rarity_good}/{rarity_total}")
    lines.append(f"{'=' * 35}")

    if pros:
        lines.append(f"\n(+) {', '.join(pros)}")
    if cons:
        lines.append(f"(-) {', '.join(cons)}")

    lines.append(f"\n{'─' * 35}")
    lines.append(f"TECHNIQUE ({tech_good}/{tech_total}) :")
    lines.extend(tech_lines)
    lines.append(f"\n{'─' * 35}")
    lines.append(f"BIOLOGIQUE ({bio_good}/{bio_total}) :")
    lines.extend(bio_lines)
    lines.append(f"\n{'─' * 35}")
    lines.append(f"RARETÉ ({rarity_good}/{rarity_total}) :")
    lines.extend(rarity_lines)

    return bio_norm, tech_norm, rarity_norm, combined_norm, "\n".join(lines)


def _apply_variant_scores(ws, scoring: dict | None = None) -> None:
    """Score each variant row and apply RGB color + comment on Import DefGen."""
    # Build col_name → col_index map
    col_map = {cell.value: cell.column for cell in ws[1]}
    defgen_col = col_map.get("Import DefGen")
    if defgen_col is None:
        return

    for row in ws.iter_rows(min_row=2):
        # Extract row values as dict
        row_data = {}
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            row_data[col_name] = cell.value

        bio, tech, rarity, _combined, comment_text = _score_variant(row_data, scoring)
        rgb = _scores_to_rgb(bio, tech, rarity)

        # Apply gradient fill to Import DefGen cell
        defgen_cell = ws.cell(row=row[0].row, column=defgen_col)
        defgen_cell.fill = PatternFill("solid", fgColor=rgb)
        c = Comment(comment_text, "vcf_filter")
        c.width = 450
        line_count = comment_text.count("\n") + 1
        c.height = max(300, line_count * 20)
        defgen_cell.comment = c


def _collect_hgvsc_across_sheets(workbook) -> dict[str, dict[str, list[int]]]:
    """Collect cross-sheet HGVSc locations.

    Returns {hgvsc_value: {sheet_name: [row_numbers]}} for values in 2+ sheets.
    """
    # hgvsc → {sheet_name: [row_numbers]}
    hgvsc_locs: dict[str, dict[str, list[int]]] = {}

    for ws_name in workbook.sheetnames:
        if ws_name == "_summary":
            continue
        ws = workbook[ws_name]
        col_idx = None
        for cell in ws[1]:
            if cell.value == "HGVSc":
                col_idx = cell.column
                break
        if col_idx is None:
            continue
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = str(row[0].value).strip() if row[0].value is not None else ""
            if val and val not in ("", "nan", "None"):
                hgvsc_locs.setdefault(val, {}).setdefault(ws_name, []).append(row[0].row)

    return {v: locs for v, locs in hgvsc_locs.items() if len(locs) > 1}


def _mark_duplicate_hgvsc(ws, hgvsc_map: dict[str, dict[str, list[int]]]) -> None:
    """Set blue font on rows whose HGVSc appears in multiple sheets.

    Also adds a comment on the HGVSc cell listing the other sheets.
    """
    col_idx = None
    for cell in ws[1]:
        if cell.value == "HGVSc":
            col_idx = cell.column
            break
    if col_idx is None:
        return

    current_sheet = ws.title

    # Map row → HGVSc value for duplicates
    dup_rows: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        val = str(row[0].value).strip() if row[0].value is not None else ""
        if val in hgvsc_map:
            dup_rows[row[0].row] = val

    if not dup_rows:
        return

    # Apply blue font to all cells in duplicate rows
    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        if row_num not in dup_rows:
            continue
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                underline=f.underline, strike=f.strike, color="0000FF",
            )

    # Add comment on HGVSc cell listing other sheets (append to existing if any)
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        row_num = row[0].row
        if row_num not in dup_rows:
            continue
        hgvsc_val = dup_rows[row_num]
        other_sheets = sorted(s for s in hgvsc_map[hgvsc_val] if s != current_sheet)
        if other_sheets:
            dup_text = f"Aussi dans : {', '.join(other_sheets)}"
        else:
            dup_text = f"Présent dans {len(hgvsc_map[hgvsc_val])} feuilles"
        existing = row[0].comment
        if existing:
            full_text = existing.text + "\n" + dup_text
        else:
            full_text = dup_text
        row[0].comment = Comment(full_text, "vcf_filter")


def _inject_usercom_formulas(
    ws,
    hgvsc_map: dict[str, dict[str, list[int]]],
    workbook,
) -> None:
    """Inject cross-sheet annotation formula in USER_COM (Excel 2010 compatible).

    For each cross-sheet duplicate, USER_COM gets a formula that concatenates
    USER_CLASS and USER_ANNOT from every other sheet where the same HGVSc appears.
    The formula is only set if USER_COM is currently empty.

    Strategy (no TEXTJOIN — Excel 2010 compatible):
      Each part = IF(OR(cls<>"",ann<>""), " | SHEET: "&cls&" "&ann, "")
      Concatenate all parts with &, then MID(..., 4, 9999) strips leading " | ".
    """
    # Find column letters in current sheet
    col_map = {}
    for cell in ws[1]:
        if cell.value in ("HGVSc", "USER_CLASS", "USER_ANNOT", "USER_COM"):
            col_map[cell.value] = get_column_letter(cell.column)

    if "HGVSc" not in col_map or "USER_COM" not in col_map:
        return

    current_sheet = ws.title
    hgvsc_letter = col_map["HGVSc"]
    usercom_letter = col_map["USER_COM"]

    # Build column letter maps for all other sheets
    other_col_maps: dict[str, dict[str, str]] = {}
    for ws_name in workbook.sheetnames:
        if ws_name in ("_summary", current_sheet):
            continue
        other_ws = workbook[ws_name]
        m = {}
        for cell in other_ws[1]:
            if cell.value in ("HGVSc", "USER_CLASS", "USER_ANNOT"):
                m[cell.value] = get_column_letter(cell.column)
        if "HGVSc" in m:
            other_col_maps[ws_name] = m

    # For each duplicate row, build the formula
    for row in ws.iter_rows(min_row=2, min_col=ws[hgvsc_letter + "1"].column,
                            max_col=ws[hgvsc_letter + "1"].column):
        cell = row[0]
        val = str(cell.value).strip() if cell.value is not None else ""
        if val not in hgvsc_map:
            continue

        row_num = cell.row
        usercom_cell = ws[f"{usercom_letter}{row_num}"]

        # Only inject if USER_COM is empty
        existing = usercom_cell.value
        if existing is not None and str(existing).strip() not in ("", "nan", "None"):
            continue

        # Build formula parts: one per other sheet where HGVSc appears
        parts = []
        for sheet_name in sorted(hgvsc_map[val]):
            if sheet_name == current_sheet:
                continue
            if sheet_name not in other_col_maps:
                continue
            ocm = other_col_maps[sheet_name]
            other_rows = hgvsc_map[val].get(sheet_name, [])
            if not other_rows:
                continue

            orow = other_rows[0]
            # Always quote sheet name — avoids issues with underscores,
            # digits at start (4_CALLERS), or other edge cases
            safe_name = f"'{sheet_name}'"

            cls_ref = f"{safe_name}!{ocm.get('USER_CLASS', 'A')}{orow}"
            ann_ref = f"{safe_name}!{ocm.get('USER_ANNOT', 'A')}{orow}"

            # Each part prepends " | " so we can strip the first one with MID
            part = (
                f'IF(OR({cls_ref}<>"",{ann_ref}<>""),'
                f'" | {sheet_name}: "&{cls_ref}&" "&{ann_ref},"")'
            )
            parts.append(part)

        if parts:
            # Concatenate all parts, then strip leading " | " (3 chars) with MID
            concat = "&".join(parts)
            formula = f"=MID({concat},4,9999)"
            usercom_cell.value = formula


def _in_neighbor_gene(chrom: str, pos) -> str | None:
    """Return neighbor gene name if (chrom, pos) falls within a known neighbor region."""
    chrom = str(chrom).strip()
    if not chrom.startswith("chr"):
        chrom = f"chr{chrom}"
    try:
        pos = int(pos)
    except (TypeError, ValueError):
        return None
    for gene, coords in _NEIGHBOR_GENES.items():
        if chrom == coords["chr"] and coords["start"] <= pos <= coords["end"]:
            return gene
    return None


def _mark_downstream_neighbor(ws) -> None:
    """Blue-fill Consequence cells for downstream_gene_variant landing in neighbor genes."""
    col_idx = {cell.value: cell.column for cell in ws[1]}
    csq_col = col_idx.get("Consequence")
    chr_col = col_idx.get("CHR")
    pos_col = col_idx.get("POS")
    if not all((csq_col, chr_col, pos_col)):
        return

    for row in ws.iter_rows(min_row=2):
        csq_val = str(ws.cell(row=row[0].row, column=csq_col).value or "")
        if "downstream_gene_variant" not in csq_val:
            continue
        chrom = ws.cell(row=row[0].row, column=chr_col).value
        pos = ws.cell(row=row[0].row, column=pos_col).value
        gene = _in_neighbor_gene(chrom, pos)
        if gene:
            cell = ws.cell(row=row[0].row, column=csq_col)
            cell.fill = _BLUE_FILL
            cell.font = _BLUE_FONT
            if not cell.comment:
                cell.comment = Comment(f"Position dans {gene}", "vcf_filter")


def _mark_hgvsc_in_extra_sheet(ws) -> None:
    """Blue-font HGVSc in extra_genes sheet for downstream variants landing in neighbor genes."""
    col_idx = {cell.value: cell.column for cell in ws[1]}
    hgvsc_col = col_idx.get("HGVSc")
    csq_col = col_idx.get("Consequence")
    chr_col = col_idx.get("CHR")
    pos_col = col_idx.get("POS")
    if not all((hgvsc_col, csq_col, chr_col, pos_col)):
        return

    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        csq_val = str(ws.cell(row=r, column=csq_col).value or "")
        if "downstream_gene_variant" not in csq_val:
            continue
        chrom = ws.cell(row=r, column=chr_col).value
        pos = ws.cell(row=r, column=pos_col).value
        gene = _in_neighbor_gene(chrom, pos)
        if gene:
            hcell = ws.cell(row=r, column=hgvsc_col)
            hcell.font = _BLUE_FONT


_C_RED_FONT  = Font(color="9C0006", size=10, bold=True)   # rouge foncé texte

# Seuils par défaut pour la discordance HTZ/mosaïque — surchargeables dans config.yaml
_DEFAULT_DISCORDANCE = {
    "htz_recurrent_in_mosaic": {
        "enabled": True,
        "min_rncnt_htz": 3,      # RNCNT_HTZ ≥ X dans le run → pattern HTZ habituel
        "mosaic_af_max": 35.0,   # AF < 35 % ici → ici en mosaïque
    },
    "mosaic_recurrent_in_htz": {
        "enabled": True,
        "min_rncnt_low": 3,      # RNCNT_LOW ≥ X dans le run → pattern mosaïque habituel
        "htz_af_min": 35.0,      # AF ≥ 35 % ici → ici en hétérozygote
        "htz_af_max": 65.0,
    },
}


def _mark_af_pattern_discordance(ws, discordance_cfg: dict | None = None) -> None:
    """Red fill + comment on HGVSc when AF pattern is discordant with run cohort.

    Case A — HTZ recurrent + mosaic here:
      RNCNT_HTZ ≥ threshold, AF_PCT < mosaic_af_max
      → Variant classiquement HTZ dans le run, ici en mosaïque — artefact probable

    Case B — Mosaic recurrent + HTZ here:
      RNCNT_LOW ≥ threshold, htz_af_min ≤ AF_PCT ≤ htz_af_max
      → Variant classiquement mosaïque dans le run, ici en hétérozygote — à regarder
    """
    cfg = _deep_merge(_DEFAULT_DISCORDANCE, discordance_cfg or {})
    col_idx = {cell.value: cell.column for cell in ws[1]}

    hgvsc_col  = col_idx.get("HGVSc")
    af_col     = col_idx.get("AF_PCT")
    rncnt_htz_col = col_idx.get("RNCNT_HTZ")
    rncnt_low_col = col_idx.get("RNCNT_LOW")
    if not hgvsc_col or not af_col:
        return

    cfg_a = cfg["htz_recurrent_in_mosaic"]
    cfg_b = cfg["mosaic_recurrent_in_htz"]

    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        af = _num(ws.cell(row=r, column=af_col).value)
        if af is None:
            continue

        rncnt_htz = _num(ws.cell(row=r, column=rncnt_htz_col).value) if rncnt_htz_col else None
        rncnt_low = _num(ws.cell(row=r, column=rncnt_low_col).value) if rncnt_low_col else None

        msg = None

        # Case A : HTZ récurrent dans le run, mais ici en mosaïque
        if (cfg_a.get("enabled", True)
                and rncnt_htz is not None
                and rncnt_htz >= cfg_a["min_rncnt_htz"]
                and af < cfg_a["mosaic_af_max"]):
            msg = (
                f"⚠ ATTENTION — Discordance AF/run\n"
                f"Variant récurrent en HTZ dans le run (RNCNT_HTZ={int(rncnt_htz)})\n"
                f"mais ici en mosaïque (AF={af:.1f}%)\n"
                f"→ Probable artefact séquentiel ou événement somatique rare"
            )

        # Case B : mosaïque récurrent dans le run, mais ici en hétérozygote
        elif (cfg_b.get("enabled", True)
                and rncnt_low is not None
                and rncnt_low >= cfg_b["min_rncnt_low"]
                and cfg_b["htz_af_min"] <= af <= cfg_b["htz_af_max"]):
            msg = (
                f"⚠ ATTENTION — Discordance AF/run\n"
                f"Variant récurrent en mosaïque dans le run (RNCNT_LOW={int(rncnt_low)})\n"
                f"mais ici en hétérozygote (AF={af:.1f}%)\n"
                f"→ À vérifier sur IGV : potentiellement pathogène ou artefact"
            )

        if msg:
            cell = ws.cell(row=r, column=hgvsc_col)
            cell.fill = _C_RED
            cell.font = _C_RED_FONT
            c = Comment(msg, "vcf_filter")
            c.width = 380
            c.height = 100
            cell.comment = c


def _apply_cell_colors(ws) -> None:
    """Apply conditional fills to data cells (rows 2+) based on column rules."""
    # Build col_name → col_index map from header row
    col_idx = {cell.value: cell.column for cell in ws[1]}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            fill = _cell_fill(col_name, cell.value)
            if fill is not None:
                cell.fill = fill


def _format_sheet(ws) -> None:
    """Apply NiourK-style formatting: header, auto-width, fixed height."""
    # Header row
    for cell in ws[1]:
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # Auto column width (max content, capped at 50)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)

    # Fixed row height + no wrap for data rows
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[i].height = _ROW_HEIGHT
        for cell in row:
            cell.font      = _DATA_FONT
            cell.alignment = _NO_WRAP

    # Freeze header row
    ws.freeze_panes = "A2"


def _fix_lovd_url(df: pd.DataFrame) -> pd.DataFrame:
    """Replace obsolete LOVD2 search URLs with LOVD3 gene page URLs.

    NiourK/VEP generates URLs like:
      http://www.lovd.nl/search.php?build=hg38&position=chr16:2083640_2083640
    These are LOVD2 format and no longer resolve.

    Replacement: https://databases.lovd.nl/shared/genes/<SYMBOL>
    which is the active LOVD3 gene page for that gene.
    """
    if "LOVD" not in df.columns or "SYMBOL" not in df.columns:
        return df
    df = df.copy()
    df["LOVD"] = df["LOVD"].astype(object)  # évite LossySetitemError si colonne float64/NaN
    old_pattern = r"^https?://www\.lovd\.nl/search\.php"
    is_old = df["LOVD"].astype(str).str.match(old_pattern)
    df.loc[is_old, "LOVD"] = df.loc[is_old, "SYMBOL"].apply(
        lambda s: f"https://databases.lovd.nl/shared/genes/{s}"
        if s and str(s) not in ("", "nan", "None")
        else ""
    )
    return df


def write_output(
    output_path: Path,
    tiers: dict[str, pd.DataFrame],
    pre_report: str,
    config_path: str,
    output_columns: list[str],
    scoring: dict | None = None,
    discordance: dict | None = None,
) -> None:
    """Write filtered sheets to a formatted multi-sheet Excel file."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # _summary sheet
        cfg_hash = hashlib.md5(Path(config_path).read_bytes()).hexdigest()[:8]
        summary_rows = (
            [{"Key": "config_path", "Value": config_path},
             {"Key": "config_md5",  "Value": cfg_hash},
             {"Key": "date",        "Value": date.today().isoformat()},
             {"Key": "",            "Value": ""}]
            + [{"Key": "", "Value": line} for line in pre_report.split("\n")]
        )
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="_summary", index=False)
        _format_sheet(writer.sheets["_summary"])

        # Data sheets — write all first, then apply cross-sheet formatting
        for name, tier_df in tiers.items():
            cols = [c for c in output_columns if c in tier_df.columns]
            _fix_lovd_url(tier_df)[cols].to_excel(writer, sheet_name=name, index=False)
            _format_sheet(writer.sheets[name])
            _apply_cell_colors(writer.sheets[name])
            _apply_variant_scores(writer.sheets[name], scoring)
            _mark_af_pattern_discordance(writer.sheets[name], discordance)

        # Cross-sheet duplicate detection: collect HGVSc across all sheets
        dup_hgvsc = _collect_hgvsc_across_sheets(writer.book)
        for name in tiers:
            _mark_duplicate_hgvsc(writer.sheets[name], dup_hgvsc)
            _inject_usercom_formulas(writer.sheets[name], dup_hgvsc, writer.book)


# ---------------------------------------------------------------------------
# 6. CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Filter NiourK XLSX for TSC1/TSC2 pathogenic candidates.")
    parser.add_argument("xlsx", type=Path, help="Input NiourK XLSX file")
    parser.add_argument("--config", type=Path, default=Path(__file__).parent / "config.yaml", help="YAML config file")
    parser.add_argument("--beta", action="store_true", help="Skip pre-exclusions — apply sheet filters on raw variants only")
    args = parser.parse_args(argv)

    cfg = load_config(args.config)

    # Load + gene filter + REF_TRANS dedup
    df = load_variants(args.xlsx, genes=cfg.get("genes"))
    total_rows = len(df)
    unique_variants = len(df.drop_duplicates(subset=["CHR", "POS", "REF", "ALT"]))

    # Exclusions (skipped in beta mode)
    if args.beta:
        df_filtered = df
        exc_report = {}
        print("[BETA] Pré-exclusions désactivées — filtres feuilles uniquement")
    else:
        df_filtered, exc_report = apply_exclusions(df, cfg)

    # Sheets (fiche d'analyse STB)
    sheets = classify_sheets(df_filtered, cfg)

    # Feuille fourre-tout pour les gènes extra (RHEB, PKD1, …) — aucun filtre qualité
    extra_genes = cfg.get("sheets", {}).get("extra_genes", [])
    if extra_genes:
        df_extra = load_variants(args.xlsx, genes=extra_genes)
        if not df_extra.empty:
            sheet_name = "_".join(extra_genes)
            sheets[sheet_name] = df_extra

    # Pre-report
    input_name = args.xlsx.stem
    report = build_pre_report(total_rows, unique_variants, exc_report, sheets, str(args.config), input_name)
    print(report)

    # Write output
    output_path = args.xlsx.with_name(f"{args.xlsx.stem}_filtered.xlsx")
    write_output(output_path, sheets, report, str(args.config), cfg.get("output_columns", []),
                 scoring=cfg.get("scoring"),
                 discordance=cfg.get("af_discordance"))
    print(f"\n→ Excel écrit : {output_path}")


if __name__ == "__main__":
    main()
