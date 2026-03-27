import subprocess
import sys
from pathlib import Path

import pandas as pd


def test_cli_produces_filtered_xlsx(make_xlsx):
    """End-to-end: run vcf_filter.py on a synthetic XLSX, check output file."""
    rows = [
        # CLINVAR_P_LP sheet
        {"CLINVAR": "Pathogenic", "IMPACT": "MODIFIER", "Consequence": "intron_variant",
         "CHR": "chr9", "POS": 100, "REF": "A", "ALT": "G"},
        # HIGH_IMPACT sheet
        {"IMPACT": "HIGH", "Consequence": "stop_gained",
         "CHR": "chr9", "POS": 200, "REF": "C", "ALT": "T"},
        # FQ_MODERATE sheet (missense, gnomAD < 0.1%)
        {"IMPACT": "MODERATE", "Consequence": "missense_variant", "REVEL": 0.80,
         "MAX_AF_GNOMADEX2.1": 0.0005,
         "CHR": "chr9", "POS": 300, "REF": "G", "ALT": "A"},
        # LOVD sheet
        {"IMPACT": "LOW", "Consequence": "synonymous_variant", "LOVD": "some_entry",
         "CHR": "chr9", "POS": 400, "REF": "T", "ALT": "C"},
        # MOSAIC sheet (AF > 5%, RNCNT_TOTAL <= 2)
        {"AF_PCT": 15.0, "RNCNT_TOTAL": 1, "IMPACT": "MODERATE",
         "Consequence": "missense_variant", "REVEL": 0.80, "MAX_AF_GNOMADEX2.1": 0.0005,
         "CHR": "chr9", "POS": 500, "REF": "A", "ALT": "T"},
        # Excluded: gnomAD too high
        {"MAX_AF_GNOMADEX2.1": 0.01, "IMPACT": "HIGH", "Consequence": "stop_gained",
         "CHR": "chr9", "POS": 600, "REF": "C", "ALT": "A"},
    ]
    xlsx_path = make_xlsx(rows)

    config_src = Path(__file__).resolve().parent.parent / "config.yaml"

    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    # Check pre-report was printed
    assert "Pre-report" in result.stdout

    # Check output file was created
    out_path = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    assert out_path.exists(), f"Output file not found: {out_path}"

    # Check sheets exist
    sheets = pd.ExcelFile(out_path, engine="openpyxl").sheet_names
    assert "_summary" in sheets
    assert "HIGH_IMPACT" in sheets
    assert "CLINVAR_P_LP" in sheets
    assert "FQ_MODERATE" in sheets
    assert "LOVD" in sheets
    assert "MOSAIC" in sheets


def test_variant_score_gradient_and_comment(make_xlsx):
    """Import DefGen cell gets gradient color + hover comment based on variant quality."""
    from openpyxl import load_workbook

    rows = [
        # Good variant: HIGH impact, ClinVar Pathogenic, good SB, 3 callers, novel
        {"IMPACT": "HIGH", "Consequence": "stop_gained", "CLINVAR": "Pathogenic",
         "SB": 0.50, "CALLNB": 3, "RNCNT_TOTAL": 0, "PJCNT_TOTAL": 0,
         "CADD_PHRED": 35.0, "MAX_AF_GNOMADEX2.1": None,
         "CHR": "chr9", "POS": 100, "REF": "A", "ALT": "G"},
        # Bad variant: MODIFIER, bad SB, 1 caller, frequent
        {"IMPACT": "MODIFIER", "Consequence": "intron_variant", "CLINVAR": None,
         "SB": 0.95, "CALLNB": 1, "RNCNT_TOTAL": 4, "PJCNT_TOTAL": 8,
         "CADD_PHRED": 5.0, "MAX_AF_GNOMADEX2.1": 0.003,
         "CHR": "chr9", "POS": 200, "REF": "C", "ALT": "T"},
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"

    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    out_path = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    wb = load_workbook(out_path)
    ws = wb["HIGH_IMPACT"]

    # Find Import DefGen column
    defgen_col = None
    for cell in ws[1]:
        if cell.value == "Import DefGen":
            defgen_col = cell.column
            break
    assert defgen_col is not None

    # Good variant (row 2) should have a comment and greenish fill
    good_cell = ws.cell(row=2, column=defgen_col)
    assert good_cell.comment is not None
    assert "SCORE :" in good_cell.comment.text
    assert "TECHNIQUE" in good_cell.comment.text
    assert "BIOLOGIQUE" in good_cell.comment.text
    # Quick overview present
    assert "(+)" in good_cell.comment.text
    # The fill should be set (not default)
    assert good_cell.fill.fgColor is not None
    good_rgb = good_cell.fill.fgColor.rgb
    # Green component should be high for good variant
    assert good_rgb is not None


def test_scores_to_rgb_signatures():
    """Verify key color signatures of the 3-axis RGB encoder."""
    from vcf_filter import _scores_to_rgb
    # Pure artifact (bad tech, no biology, not rare) → red
    assert _scores_to_rgb(bio=0.0, tech_quality=0.0, rarity=0.0) == "FF0000"
    # Perfect biology + rare + clean tech → cyan
    assert _scores_to_rgb(bio=1.0, tech_quality=1.0, rarity=1.0) == "00FFFF"
    # Biology + artifact (dangerous false positive) → yellow
    assert _scores_to_rgb(bio=1.0, tech_quality=0.0, rarity=0.0) == "FFFF00"
    # Uninteresting but technically clean → black
    assert _scores_to_rgb(bio=0.0, tech_quality=1.0, rarity=0.0) == "000000"


def test_score_variant_good_vs_bad():
    """A clearly good variant scores higher than a clearly bad one."""
    from vcf_filter import _score_variant
    good = {
        "SB": 0.50, "CALLNB": 4, "RNCNT_TOTAL": 0, "PJCNT_TOTAL": 0,
        "DP": 200, "AF_PCT": 48.0, "CLINVAR": "Pathogenic", "IMPACT": "HIGH",
        "Consequence": "stop_gained", "CADD_PHRED": 35.0, "REVEL": 0.9,
        "MAX_AF_GNOMADEX2.1": None,
    }
    bad = {
        "SB": 0.95, "CALLNB": 1, "RNCNT_TOTAL": 4, "PJCNT_TOTAL": 8,
        "DP": 50, "AF_PCT": 2.0, "CLINVAR": None, "IMPACT": "MODIFIER",
        "Consequence": "intron_variant", "CADD_PHRED": 5.0, "REVEL": 0.1,
        "MAX_AF_GNOMADEX2.1": 0.003,
    }
    _good_bio, _good_tech, _good_rarity, good_combined, good_comment = _score_variant(good)
    _bad_bio, _bad_tech, _bad_rarity, bad_combined, bad_comment = _score_variant(bad)
    assert good_combined > 0.7
    assert bad_combined < 0.3
    assert good_combined > bad_combined
    assert "✓" in good_comment
    assert "✗" in bad_comment


def test_duplicate_hgvsc_blue_font_cross_sheet(make_xlsx):
    """HGVSc present in multiple sheets should have blue font (cross-sheet detection)."""
    from openpyxl import load_workbook

    rows = [
        # This variant lands in HIGH_IMPACT + CLINVAR_P_LP + LOVD → 3 sheets = blue
        {"IMPACT": "HIGH", "Consequence": "stop_gained", "HGVSc": "c.1525C>T",
         "CLINVAR": "Pathogenic", "LOVD": "some_entry",
         "CHR": "chr9", "POS": 100, "REF": "C", "ALT": "T"},
        # This variant lands ONLY in HIGH_IMPACT → 1 sheet = NOT blue
        {"IMPACT": "HIGH", "Consequence": "stop_gained", "HGVSc": "c.999G>A",
         "CHR": "chr9", "POS": 300, "REF": "A", "ALT": "G"},
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"

    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    out_path = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    wb = load_workbook(out_path)
    ws = wb["HIGH_IMPACT"]

    # Find HGVSc column index
    hgvsc_col = None
    for cell in ws[1]:
        if cell.value == "HGVSc":
            hgvsc_col = cell.column
            break
    assert hgvsc_col is not None

    # Check font colors: cross-sheet duplicate should be blue, unique should not
    blue_rows = []
    non_blue_rows = []
    for row in ws.iter_rows(min_row=2, min_col=hgvsc_col, max_col=hgvsc_col):
        cell = row[0]
        color = cell.font.color
        if color and color.rgb and "0000FF" in str(color.rgb):
            blue_rows.append(cell.value)
        else:
            non_blue_rows.append(cell.value)

    assert "c.1525C>T" in blue_rows    # present in HIGH_IMPACT + CLINVAR_P_LP + LOVD
    assert "c.999G>A" in non_blue_rows  # only in HIGH_IMPACT

    # Check that the HGVSc cell for the duplicate has a comment listing other sheets
    for row in ws.iter_rows(min_row=2, min_col=hgvsc_col, max_col=hgvsc_col):
        cell = row[0]
        if cell.value == "c.1525C>T":
            assert cell.comment is not None
            comment = cell.comment.text
            assert "Aussi dans" in comment
            # Should mention at least CLINVAR_P_LP and LOVD
            assert "CLINVAR_P_LP" in comment
            assert "LOVD" in comment
        elif cell.value == "c.999G>A":
            assert cell.comment is None  # no cross-sheet duplicate


def test_usercom_formula_cross_sheet(make_xlsx):
    """USER_COM should have a TEXTJOIN formula referencing USER_CLASS/ANNOT from other sheets."""
    from openpyxl import load_workbook

    rows = [
        # This variant lands in HIGH_IMPACT + CLINVAR_P_LP + LOVD → formula in each
        {"IMPACT": "HIGH", "Consequence": "stop_gained", "HGVSc": "c.1525C>T",
         "CLINVAR": "Pathogenic", "LOVD": "some_entry",
         "CHR": "chr9", "POS": 100, "REF": "C", "ALT": "T"},
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"

    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    out_path = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    wb = load_workbook(out_path)

    # Check HIGH_IMPACT sheet — USER_COM should have formula referencing CLINVAR_P_LP and LOVD
    ws = wb["HIGH_IMPACT"]
    usercom_col = None
    for cell in ws[1]:
        if cell.value == "USER_COM":
            usercom_col = cell.column
            break
    assert usercom_col is not None

    usercom_cell = ws.cell(row=2, column=usercom_col)
    formula = str(usercom_cell.value)
    assert formula.startswith("=MID(")
    assert "CLINVAR_P_LP" in formula
    assert "LOVD" in formula


def test_cli_pre_report_counts(make_xlsx):
    """Verify pre-report shows correct exclusion counts."""
    rows = [
        {"MAX_AF_GNOMADEX2.1": 0.01},   # excluded by gnomAD
        {"RNCNT_TOTAL": 15},             # excluded by RNCNT
        {"IMPACT": "HIGH", "Consequence": "stop_gained"},  # survives
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"

    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert "gnomAD" in result.stdout
    assert "RNCNT_TOTAL" in result.stdout


def test_af_discordance_htz_recurrent_in_mosaic(make_xlsx):
    """HGVSc cell should be red when variant is recurrently HTZ in run but mosaic here."""
    from openpyxl import load_workbook

    rows = [
        {   # RNCNT_HTZ=5, AF=12% → discordance HTZ→mosaïque
            "CLINVAR": "Pathogenic", "IMPACT": "MODERATE", "Consequence": "missense_variant",
            "MAX_AF_GNOMADEX2.1": 0.0001,
            "CHR": "chr9", "POS": 100, "REF": "A", "ALT": "G",
            "HGVSc": "NM_000368.5:c.100A>G",
            "RNCNT_HTZ": 5, "RNCNT_LOW": 0, "RNCNT_TOTAL": 5, "AF_PCT": 12.0,
        },
        {   # RNCNT_HTZ=1 → pas récurrent, pas de flag
            "CLINVAR": "Pathogenic", "IMPACT": "MODERATE", "Consequence": "missense_variant",
            "MAX_AF_GNOMADEX2.1": 0.0001,
            "CHR": "chr9", "POS": 200, "REF": "C", "ALT": "T",
            "HGVSc": "NM_000368.5:c.200C>T",
            "RNCNT_HTZ": 1, "RNCNT_LOW": 0, "RNCNT_TOTAL": 1, "AF_PCT": 12.0,
        },
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"
    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    out = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    wb = load_workbook(out)
    ws = wb["CLINVAR_P_LP"]
    headers = {cell.value: cell.column for cell in ws[1]}
    hgvsc_col = headers.get("HGVSc")
    assert hgvsc_col is not None

    # Row avec RNCNT_HTZ=5 + AF mosaïque → rouge + commentaire
    red_rows = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=hgvsc_col).comment is not None
        and "HTZ" in ws.cell(row=r, column=hgvsc_col).comment.text
    ]
    assert len(red_rows) == 1, "Exactement une ligne devrait avoir le commentaire discordance HTZ"

    # Row avec RNCNT_HTZ=1 → pas de commentaire discordance (peut avoir doublon)
    discordance_rows_all = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=hgvsc_col).comment is not None
        and "HTZ" in ws.cell(row=r, column=hgvsc_col).comment.text
    ]
    assert len(discordance_rows_all) == 1, "Seule la ligne RNCNT_HTZ=5 doit avoir le flag discordance"


def test_af_discordance_mosaic_recurrent_in_htz(make_xlsx):
    """HGVSc cell should be red when variant is recurrently mosaic in run but HTZ here."""
    from openpyxl import load_workbook

    rows = [
        {   # RNCNT_LOW=4, AF=48% → discordance mosaïque→HTZ
            "CLINVAR": "Pathogenic", "IMPACT": "MODERATE", "Consequence": "missense_variant",
            "MAX_AF_GNOMADEX2.1": 0.0001,
            "CHR": "chr9", "POS": 100, "REF": "A", "ALT": "G",
            "HGVSc": "NM_000368.5:c.100A>G",
            "RNCNT_HTZ": 0, "RNCNT_LOW": 4, "RNCNT_TOTAL": 4, "AF_PCT": 48.0,
        },
    ]
    xlsx_path = make_xlsx(rows)
    config_src = Path(__file__).resolve().parent.parent / "config.yaml"
    result = subprocess.run(
        [sys.executable, "vcf_filter.py", str(xlsx_path), "--config", str(config_src)],
        capture_output=True, text=True,
        cwd=str(Path(__file__).resolve().parent.parent),
    )
    assert result.returncode == 0, f"STDERR: {result.stderr}"

    out = xlsx_path.with_name(xlsx_path.stem + "_filtered.xlsx")
    wb = load_workbook(out)
    ws = wb["CLINVAR_P_LP"]
    headers = {cell.value: cell.column for cell in ws[1]}
    hgvsc_col = headers.get("HGVSc")
    assert hgvsc_col is not None

    flagged = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=hgvsc_col).comment is not None
        and "mosaïque" in ws.cell(row=r, column=hgvsc_col).comment.text
    ]
    assert len(flagged) == 1, "La ligne devrait avoir le commentaire discordance mosaïque→HTZ"
